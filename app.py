import os
import json
import pandas as pd
import streamlit as st
import google.generativeai as genai
from dotenv import load_dotenv
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

load_dotenv()

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel("gemini-2.5-flash")


def generate_content(business_type, topic):
    prompt = f"""
You are an expert social media content creator.

Create content for:

Business Type: {business_type}
Topic: {topic}

Return ONLY valid JSON.
Do not use markdown.
Do not add explanation.

JSON format:
{{
  "instagram_caption": "",
  "reel_script": "",
  "facebook_post": "",
  "linkedin_post": "",
  "youtube_title": "",
  "youtube_description": "",
  "hashtags": ""
}}
"""

    response = model.generate_content(prompt)
    text = response.text.strip()
    text = text.replace("```json", "").replace("```", "").strip()
    return json.loads(text)


def create_excel_file(result, business_type, topic):
    data = {
        "Date": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        "Business Type": [business_type],
        "Topic": [topic],
        "Instagram Caption": [result["instagram_caption"]],
        "Reel Script": [result["reel_script"]],
        "Facebook Post": [result["facebook_post"]],
        "LinkedIn Post": [result["linkedin_post"]],
        "YouTube Title": [result["youtube_title"]],
        "YouTube Description": [result["youtube_description"]],
        "Hashtags": [result["hashtags"]]
    }

    df = pd.DataFrame(data)

    file_name = "generated_posts_clean.xlsx"
    df.to_excel(file_name, index=False)

    wb = load_workbook(file_name)
    ws = wb.active

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 22,
        "B": 18,
        "C": 25,
        "D": 45,
        "E": 55,
        "F": 55,
        "G": 55,
        "H": 35,
        "I": 55,
        "J": 45,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 180

    ws.freeze_panes = "A2"
    wb.save(file_name)

    return file_name


st.set_page_config(
    page_title="AI Social Media Content Agent",
    page_icon="🚀",
    layout="wide"
)

st.title("🚀 AI Social Media Content Agent")
st.write("Generate Instagram, Facebook, LinkedIn and YouTube content using AI.")

col1, col2 = st.columns(2)

with col1:
    business_type = st.text_input(
        "Business Type",
        placeholder="Example: Real Estate Agent, Gym, Dental Clinic, AI Agency..."
    )

with col2:
    topic = st.text_input(
        "Enter Topic",
        placeholder="Example: 2 BHK Flats in Navi Mumbai"
    )

if st.button("Generate Content", use_container_width=True):

    if not business_type:
        st.warning("Please enter business type.")

    elif not topic:
        st.warning("Please enter a topic.")

    elif not GEMINI_API_KEY:
        st.error("Gemini API key missing. Check your .env file.")

    else:
        with st.spinner("Generating content..."):

            try:
                result = generate_content(business_type, topic)

                st.success("Content Generated!")

                tab1, tab2, tab3, tab4, tab5 = st.tabs(
                    ["Instagram", "Reel Script", "Facebook", "LinkedIn", "YouTube"]
                )

                with tab1:
                    st.subheader("Instagram Caption")
                    st.text_area("Copy Instagram Caption", result["instagram_caption"], height=220)

                    st.subheader("Hashtags")
                    st.text_area("Copy Hashtags", result["hashtags"], height=120)

                with tab2:
                    st.subheader("Instagram Reel Script")
                    st.text_area("Copy Reel Script", result["reel_script"], height=350)

                with tab3:
                    st.subheader("Facebook Post")
                    st.text_area("Copy Facebook Post", result["facebook_post"], height=300)

                with tab4:
                    st.subheader("LinkedIn Post")
                    st.text_area("Copy LinkedIn Post", result["linkedin_post"], height=300)

                with tab5:
                    st.subheader("YouTube Shorts Title")
                    st.text_area("Copy YouTube Title", result["youtube_title"], height=100)

                    st.subheader("YouTube Description")
                    st.text_area("Copy YouTube Description", result["youtube_description"], height=220)

                file_name = create_excel_file(result, business_type, topic)

                with open(file_name, "rb") as file:
                    st.download_button(
                        label="📥 Download Clean Excel",
                        data=file,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

            except Exception as e:
                st.error("Something went wrong.")
                st.write(e)